import base64
import os
import subprocess

import xlrd
import string
import json
from openpyxl import load_workbook,Workbook
import openpyxl as op

workingPath = ""


# 批量生成用来修改的html文件对应的txt文件，之后用下一个函数将.html后缀改为.txt。
#用来生成之前的记录在表格中的内容
def generateHtmlold():

    f1 = open("/home/liu/桌面/gumtree_tmp/special_html/1.txt", 'r', encoding='utf-8')
    content = f1.read()
    for i in range(50, 101):
        with open("/home/liu/桌面/gumtree_tmp/special_html/" + str(i) + ".txt", "w", encoding='utf-8') as f:
            f.write(content)

    stra = "<html>\n<body>\n<a href=\"https://github.com/"
    strb = "\" style=\"margin-left:50px;\">所在commit地址</a>\n<code>\n<pre style=\"font-size: 20px;font-family:'Times New Roman', Times, serif;color:brown;\">\n"
    strc = "\n</pre>\n</code>\n<p>message:</p>\n<p>李蓝天：改动分类：原因：</p>\n<p>刘志浩：改动分类：原因：</p>\n</body>\n</html>"

    data = xlrd.open_workbook("/home/liu/桌面/gumtree_tmp/newanalysis.xlsx")  # 读取存放分析结果的xlsx文件
    fspecial = open("/home/liu/PycharmProjects/pythonProject2/special1.txt", "r", encoding='utf-8')  # 打开存放commit连接的文件
    s1 = fspecial.readlines()  # 按行读取存放commit连接的文件，每一行有一个连接
    table = data.sheets()[6]  # 读取第六张表的数据
    rows = table.nrows  # 获取行数
    cols = table.col_values(0)  # 读取第一列的内容
    ###从第一行读到最后一行，读取第二列的数据
    for i in range(0, rows):
        cols[i] = int(cols[i])  # 数组中的内容原本是浮点数类型，转换为整型
        fpatch = open("/home/liu/PycharmProjects/GenerateAST/special/" + str(cols[i]) + ".txt", 'r',
                      encoding='utf-8')  # 读取对应的patch文件
        pacth_content = fpatch.read()  # 获得patch中的内容
        # 删除最后多余的sha，最后形式是 用户名/仓库名/commit/sha
        link1, link2, link3 = s1[cols[i]].rpartition('/')
        s1[cols[i]] = link1
        # 将各部分连接起来
        strall = stra + s1[cols[i]] + strb + pacth_content + strc
        fsave = open("/home/liu/桌面/gumtree_tmp/special_html_test/" + str(i) + ".txt", "w", encoding="utf-8")  # 存文件
        fsave.write(strall)


###### 批量修改文件名称（将.txt变为.html）
def txttohtml():
    path = "/home/liu/桌面/gumtree_tmp/special_html_test"
    os.chdir(path)
    files = os.listdir(path)

    for filename in files:
        portion = os.path.splitext(filename)  # 分离文件名与扩展名
        # 如果后缀是.txt
        if portion[1] == '.txt':
            # 重新组合文件名和后缀名
            newname = portion[0] + '.html'  # 修改为.html
            # newname = "ss.html"
            os.rename(filename, newname)

# 批量生成用来修改的html文件对应的txt文件，之后用下一个函数将.html后缀改为.txt。
def generateHtml(pathstr,fLinkstr):
    path = pathstr
    os.chdir(path)
    files = os.listdir(path)
    fLink = open(fLinkstr,"r",encoding="utf-8")
    s1 = fLink.readlines()
    stra = "<html>\n<body>\n<a href=\""
    strb = "\" style=\"margin-left:50px;\">所在commit地址</a>\n<code>\n<pre style=\"font-size: 20px;font-family:'Times New Roman', Times, serif;color:brown;\">\n"
    strc = "\n</pre>\n</code>\n<p>message:</p>\n<p>李蓝天：改动分类：原因：</p>\n<p>梁叶剑：改动分类：原因：</p>\n<p>刘志浩：改动分类：原因：</p>\n</body>\n</html>"
    os.mkdir("webpage")
    for filename in files:
        if (filename == "webpage") or (filename == "saveLink.txt"):
            continue
        strall = ""
        portion = os.path.splitext(filename)#分离文件名与扩展名
        ftmp = open(filename,"r",encoding='utf-8')
        tmpnum = int(portion[0])
        strall = stra + s1[tmpnum] + strb + ftmp.read() + strc
        savefile = open(pathstr + "/webpage/" + str(tmpnum) + ".txt",'w',encoding='utf-8')
        savefile.write(strall)


# 批量生成网址
def generateLink():
    for i in range(0, 222):
        str1 = "https://highbe.github.io/SolidityWorm/HtmlPlace/" + str(i) + ".html\n"
        print(str1)

#将当前文件夹下的文件对应的网址找到并存到一个文件中
def collectLink(pathstr,fLinkstr):
    fLink = open(fLinkstr,'r',encoding='utf-8')#存放所有地址的文件
    s1 = fLink.readlines()
    path = pathstr#分类后的文件夹
    os.chdir(path)
    files = os.listdir(path)
    files.sort(key=lambda x:int(x.split('.')[0]))#读取文件后按照文件名称排序,要求文件名要格式一致

    content = ""#存放所有需要的网址
    for filename in files:
        portion = os.path.splitext(filename)#分离文件名与扩展名,portion[0]是文件名，portion[1]是扩展名
        tmpnum = int(portion[0])
        content += portion[0] + " " + s1[tmpnum] + "\n"

    fsave = open("saveLink.txt","w",encoding="utf-8")
    fsave.write(content)


#统计单个文件中的行数
def countlines(path,count,emitcount):
    catalog = open(path,"r",encoding="utf-8",errors="ignore")
    lines = catalog.readlines()
    i = 0
    while i < len(lines):
        # line = lines[i].strip()
        lines[i] = lines[i].strip()
        # 如果是空行直接跳过
        if lines[i] == "":
            i += 1
            continue
        #遇到emit则emit计数器与总计数器都加一，同时直接略过下面的判断语句
        if lines[i].startswith("emit "):
            emitcount[0] += 1
            count[0] += 1
            i += 1
            continue

        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//"):
            i += 1
            continue
        if lines[i].startswith("/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):#防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue

        if lines[i].find(";") != -1 or lines[i].find("{") != -1:
            count[0] += 1
            i += 1
        else :
            i += 1


# 每个项目需要单独调用该函数  找到每个项目中的所有solidity文件
def traversal(path,count,emitcount):
    # 首先遍历当前目录所有文件及文件夹
    file_list = os.listdir(path)
    # 循环判断每个元素是否是文件夹还是文件，是文件夹的话，递归
    for file in file_list:
        # 利用os.path.join()方法取得路径全名，并存入cur_path变量，否则每次只能遍历一层目录
        cur_path = os.path.join(path, file)
        # 判断是否是文件夹
        if os.path.isdir(cur_path):
            if os.path.islink(file) or file == 'to_outside':#判断这个文件夹是不是一个软连接，软链接可能导致死循环，只有一个项目中有软链接，名称为to_outside所以直接进行判断了
                continue
            traversal(cur_path,count,emitcount)
        # 判断是否是solidity文件
        elif cur_path.find(".sol",len(cur_path)-4,len(cur_path)) != -1:
            # print("1" + cur_path)
            countlines(cur_path,count,emitcount)
        #既不是文件夹也不是solidity文件
        else:
            continue

#统计每个项目的star数量
def countstar(path):


    file_list = os.listdir(path)
    excel_path = "/home/liu/桌面/gumtree_tmp/统计数据/emitchange汇总.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active

    for file in file_list:
        cur_path = os.path.join(path,file)
        content = open(cur_path,"r")
        strings = content.read()
        jsonStr = json.loads(strings)
        jsonStr['stargazers_count'] #收藏数
        namebefore = jsonStr['full_name']#用户名 + “/” + 仓库名 需要将两个名之间的“/”改为空格才能再表格中找到对因的文件
        nameafter = namebefore.replace('/',' ')
        print(file + " " + nameafter)
        for i in range(2,2950):
            if nameafter == ws.cell(i,1).value:# excel表格的第一行第一列是0，0
                ws.cell(i,6).value = jsonStr['stargazers_count']
                print(i)
                continue

    wb.save(excel_path)

#统计含有emit改变的commit占所有commit的比例
def changepercent(path):
    all_dir = os.listdir(path)
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(['项目名','总改动次数','emit改动次数','emit改动次数/总改动次数'])
    for dir in all_dir:
        count = 0#统计一共有多少次改动
        changecount = 0#统计一共有多少次emit改动
        tmp_path = os.path.join(path,dir)
        if os.path.isdir(tmp_path):#进入项目
            cur_path = os.path.join(path,dir)
            file_list = os.listdir(cur_path)
            for file in file_list:
                count += 1#每有一个文件就有一次改动
                print(1)
                if(isEmitChange(tmp_path+"/"+file)):#有emit的改动
                    changecount += 1

        d = dir,count,changecount,((changecount/count) if (count!=0) else -1)
        ws.append(d)
    #生成的表格的存放位置
    wb.save("/home/liu/PycharmProjects/SolidityWorm/emitchange.xlsx")


#判断是否包含emit的改变
def isEmitChange(path):
    content = open(path, "r")
    strings = content.read()
    jsonStr = json.loads(strings)
    #如果文件中没有包含一下内容说明emit不会有修改
    if ('files' not in jsonStr):
        return False
    for i in range(len(jsonStr['files'])):
        if 'patch' not in jsonStr['files'][i]:
            continue
        tmpStr = jsonStr['files'][i]['patch'].replace(' ', '')
        lines = tmpStr.split('\n')#将文件修改的patch信息用列表lines来存储

        #开始遍历每一行
        i = 0
        while i < len(lines):
            # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
            if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
                i += 1
                continue
            if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
                rightzhushi = lines[i].find("*/")
                # 找到另一半注释
                while rightzhushi == -1:
                    i += 1
                    if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                        break
                    rightzhushi = lines[i].find("*/")
                i += 1
                continue
            #如果修改的地方有emit，则返回True
            if lines[i].startswith("+emit") or lines[i].startswith("-emit"):
                return True
            #什么都没有发生 到下一行
            i += 1
    # 遍历完所有内容都没有找到修改emit
    return False

#判断是否是相同emit的改变
def samechangeamount(path):
    content = open(path, "r")
    strings = content.read()
    jsonStr = json.loads(strings)
    # 如果文件中没有包含一下内容说明emit不会有修改
    if ('files' not in jsonStr):
        return False
    for i in range(len(jsonStr['files'])):
        if 'patch' not in jsonStr['files'][i]:
            continue
        tmpStr = jsonStr['files'][i]['patch'].replace(' ', '')
        lines = tmpStr.split('\n')  # 将文件修改的patch信息用列表lines来存储

        hashtable = dict()

        # 开始遍历每一行
        i = 0
        while i < len(lines):
            # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
            if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
                i += 1
                continue
            if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
                rightzhushi = lines[i].find("*/")
                # 找到另一半注释
                while rightzhushi == -1:
                    i += 1
                    if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                        break
                    rightzhushi = lines[i].find("*/")
                i += 1
                continue
            # 如果修改的地方有emit，将修改内容存储下来，之后再遇到emit变化与之前存储内容进行比对
            if lines[i].startswith("+emit") or lines[i].startswith("-emit"):
                emitStr = lines[i].split(";")[0]#找到完整的emit语句，防止在句子末尾有注释
                if emitStr in hashtable: #找到相同的emit操作
                    return True
                else: #没有找到相同的emit操作，将现在的操作记录
                    hashtable[emitStr] = 1
            # 什么都没有发生 到下一行
            i += 1
    # 遍历完所有内容都没有找到修改emit
    return False

#统计相同emit改变占所有emit改变的比例
def samechangepercent(path):
    all_dir = os.listdir(path)
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(['项目名',  'emit改动次数','相同emit改动次数', '相同emit改动次数/emit改动次数'])
    for dir in all_dir:
        samecount = 0  # 统计一共有多少次改动
        changecount = 0  # 统计一共有多少次emit改动
        tmp_path = os.path.join(path, dir)
        if os.path.isdir(tmp_path):  # 进入项目
            cur_path = os.path.join(path, dir)
            file_list = os.listdir(cur_path)
            for file in file_list:
                print(1)
                if (isEmitChange(tmp_path + "/" + file)):  # 有emit的改动
                    changecount += 1
                if samechangeamount(tmp_path + "/" + file):
                    samecount += 1


        d = dir,changecount,samecount, ((samecount / changecount) if (changecount != 0) else -1)
        ws.append(d)
    # 生成的表格的存放位置
    wb.save("/media/liu/02F8200EF81FFE93/Liu/data/sameEmitChange4.xlsx")


#利用修改后的文件和patch文件还原出修改前的文件
def rebackSrc(patchStr, dstStr):
    lines = patchStr.split("\n")
    dst = dstStr.split("\n")

    taps = []
    patchPos = []
    for i in range(0, len(lines)):
        head = lines[i].find("@@")
        if head == 0:
            rear = lines[i][head + 2:].find("@@")
            tempstr = lines[i][head + 3:rear].split(" ")[1].split(",")[0][1:]
            taps.append(tempstr)
            patchPos.append(i)
    # 进行对patch文件定位
    srcPos = 0
    tapsPos = 0
    newStream = []
    lines[len(lines) - 1] += "\n"
    while srcPos < len(dst):
        # 表示下面接patch内容
        if tapsPos >= len(taps):
            newStream.append(dst[srcPos] + "\n")
            srcPos += 1
            continue
        if srcPos == int(taps[tapsPos]) - 1:
            patchStart = patchPos[tapsPos] + 1
            patchEnd = len(lines)
            if tapsPos < len(taps) - 1:
                patchEnd = patchPos[tapsPos + 1]
            for i in range(patchStart, patchEnd):
                if lines[i] == '':
                    newStream.append(lines[i] + "\n")
                    srcPos += 1
                else:
                    if lines[i][0] == '+':
                        srcPos += 1
                        continue
                    elif lines[i][0] == '-':
                        newStream.append(lines[i][1:] + "\n")
                    else:
                        newStream.append(lines[i] + "\n")
                        srcPos += 1
            tapsPos += 1
        else:
            newStream.append(dst[srcPos] + "\n")
            srcPos += 1
    if "No newline at end of file" in newStream[-1]:
        newStream.remove(newStream[-1])

    return "".join(newStream)

#统计单个文件中的emit行数
def countEmitLine(str):
    lines = str.split("\n")
    i = 0
    emitcount = 0
    while i < len(lines):
        lines[i] = lines[i].strip()
        # 遇到emit则emit计数器与总计数器都加一，同时直接略过下面的判断语句
        if lines[i].startswith("emit "):
            emitcount += 1
            i += 1
            continue
        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//"):
            i += 1
            continue
        if lines[i].startswith("/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue
        i += 1
    return emitcount


def emitCodeChurn(path):
    all_dir = os.listdir(path)
    wb = op.Workbook()
    ws = wb['Sheet']
    ws.append(['项目名', 'add','Code Churn','all'])
    objPath = "/home/yantong/Code/CodeLine/repos/"  # 存放项目的目录
    #一些项目存放的是数据集或类似内容，不属于我们需要的，遇到这个项目直接跳过
    notneed = ['tintinweb smart-contract-sanctuary-avalanche','tintinweb smart-contract-sanctuary-fantom','renardbebe Smart-Contract-Benchmark-Suites',
               'tintinweb smart-contract-sanctuary-arbitrum','tintinweb smart-contract-sanctuary-tron','xf97 HuangGai',
               'tintinweb smart-contract-sanctuary-optimism','gasgauge gasgauge.github.io','interfinetwork audited-codes',
               'giacomofi Neural_Smart_Ponzi_Recognition','solidproof projects','SoftSec-KAIST Smartian-Artifact',
               'bokkypoobah Tokens','makerdao spells-mainnet','Messi-Q GPSCVulDetector','kupl VeriSmart-benchmarks',
               'eff-kay solidity-nicad','ethereum solidity-underhanded-contest','Dapp-Learning-DAO Dapp-Learning']
    for j in range(0,len(all_dir)): #进入每个项目
        try:
            dir = all_dir[j]
            changecount = 0  # 统计一共有多少commit
            notmp = 0 #保存一个项目中所有改动的emit的code churn的和（不包含emit新增）
            intmp = 0 #保存一个项目中所有改动的emit的code churn的和（包含emit新增）
            alltmp = 0 #保存一个项目中所有提交的普通代码code churn的和
            cur_path = os.path.join(path, dir)

            if dir in notneed:
                continue
            if os.path.isdir(cur_path):  # 进入项目
                count = [0] #统计所有的代码行数
                emitcount = [0] #统计所有的emit行数
                file_list = os.listdir(cur_path)
                for file in file_list:#file是每次commit前的文件
                    #将版本切换到sha对应版本修改前的版本
                    os.chdir(objPath + dir)
                    nextsha = file[:-5]
                    command = "git checkout " + nextsha
                    os.system(command)
                    command = "git log"
                    returnStr = os.popen(command).read()
                    returnlines = returnStr.split("\n")
                    commitcount = 0
                    sha = ""
                    for line in returnlines:
                        if line.startswith("commit"):
                            commitcount += 1
                            if  commitcount == 2:
                                sha = line[7:]
                    if commitcount == 1: #说明在这条commit之前没有提交，即这是第一次提交，不是我们需要的，直接进行遍历下一次提交
                        continue

                    #将库切换到对应版本
                    os.system("git reset --hard")
                    os.system("git clean -dfx")#强制版本转换，不会弹提醒
                    os.system("git checkout " + sha) #切换到sha对应的版本

                    traversal(objPath + dir, count, emitcount)#统计原文件中代码和emit的行数
                    changecount += 1  # 每有一个文件就有一次改动
                    # print(file)

                    file_path = os.path.join(cur_path,file)
                    content = open(file_path,"r")
                    strings = content.read()
                    jsonStr = json.loads(strings)

                    noCount = 0  # 统计不包含新增有多少次emit次修改
                    inCount = 0  # 统计包含新增有多少次emit修改
                    allcount = 0 #统计所有代码的变动次数
                    for i in range(len(jsonStr['files'])):  #求所有文件的改动数量的和
                        if 'patch' not in jsonStr['files'][i]: #该文件没有修改
                            continue
                        if jsonStr['files'][i]['filename'][-4:] != ".sol":  #如果当前文件不存在说明不是solidity文件，直接跳过
                            continue

                        noCount += noAddChange(jsonStr['files'][i])  #统计不包含emit新增的修改
                        inCount += includeAddChange(jsonStr['files'][i])  #统计包含emit新增的修改
                        allcount += allchange(jsonStr['files'][i])#统计所有代码的修改次数

                        # print(str(file) + " " + str(inCount) + " " + str(noCount) + " " + str(allcount))
                    if emitcount[0] != 0:
                        changecount += 1
                        notmp += (noCount/emitcount[0]) #这次改动中的code churn
                        intmp += (inCount/emitcount[0])
                        alltmp += (allcount/count[0])

                if changecount != 0: #commit次数不为零 所有改动中该项目的code churn
                    notmp /= changecount
                    intmp /= changecount
                    alltmp /= changecount
                else:
                    notmp = -1
                    intmp = -1
                    alltmp = -1

                # d = dir, intmp,notmp,alltmp
                print(str(j) + " " + str(intmp) + " " + str(notmp) + " " + str(alltmp))
                # ws.append(d)
                ws.cell(j+2,1).value = dir
                ws.cell(j+2,2).value = intmp
                ws.cell(j+2,3).value = notmp
                ws.cell(j+2,4).value = alltmp
                wb.save("/home/yantong/Zhihao/code_churn/codechurnLiang.xlsx")
        except:
            continue




# 统计emit修改的行数 不包含新增emit的修改
def noAddChange(jsonStr):
    # 如果文件中没有包含一下内容说明emit不会有修改
    hashtable = {'+' : 0, '-' : 0} #记录emit的修改内容，用来配对
    changecount = 0 #统计emit修改的次数


    tmpStr = jsonStr['patch'].replace(' ', '')
    lines = tmpStr.split('\n')  # 将文件修改的patch信息用列表lines来存储

    # 开始遍历每一行
    i = 0
    while i < len(lines):
        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
            i += 1
            continue
        if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue
        #统计方法类似于所有代码的code churn 的统计方法
        if lines[i].startswith("-emit"):
            changecount += 1
        i += 1
    return changecount



#统计emit修改的函数 包含新增emit的修改
def includeAddChange(jsonStr):
    # 如果文件中没有包含一下内容说明emit不会有修改
    hashtable = {'+' : 0, '-' : 0} #记录emit的修改内容，用来配对
    changecount = 0  # 统计emit修改的次数


    tmpStr = jsonStr['patch'].replace(' ', '')
    lines = tmpStr.split('\n')  # 将文件修改的patch信息用列表lines来存储

    # 开始遍历每一行
    i = 0
    while i < len(lines):
        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
            i += 1
            continue
        if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue
        # 如果修改的地方有emit，则返回True
        if lines[i].startswith("+emit") or lines[i].startswith("-emit"):
            if lines[i].startswith("+emit"):
                if lines[i].split("(")[0].replace("+","-") in hashtable:  # 如果在表中找到能够配对的emit修改，将表中的对应内容删除，同时emit修改次数加一
                    changecount += 1
                    del hashtable[lines[i].split("(")[0].replace("+","-")]
                    hashtable['-'] -= 1
                else:
                    hashtable[lines[i].split("(")[0]] = 1
                    hashtable['+'] += 1
            else:
                if lines[i].split("(")[0].replace("-","+") in hashtable:  # 如果在表中找到能够配对的emit修改，将表中的对应内容删除，同时emit修改次数加一
                    changecount += 1
                    del hashtable[lines[i].split("(")[0].replace("-", "+")]
                    hashtable['+'] -= 1
                else:
                    hashtable[lines[i].split("(")[0]] = 1
                    hashtable['-'] += 1

        # 什么都没有发生 到下一行
        i += 1
    # 返回修改次数 加 删除次数 加 新增次数
    return changecount + hashtable['-'] + hashtable['+']


#只统计删除的修改
def allchange(jsonStr):
    changecount = 0

    lines = jsonStr['patch'].split('\n')  # 将文件修改的patch信息用列表lines来存储

    # 开始遍历每一行
    i = 0
    while i < len(lines):
        # 遇到注释，如果行首是//那么直接跳过这行，如果行首是/*需要找到*/，在这期间的内容都不要
        if lines[i].startswith("//") or lines[i].startswith("+//") or lines[i].startswith("-//"):
            i += 1
            continue
        if lines[i].startswith("/*") or lines[i].startswith("+/*") or lines[i].startswith("-/*"):
            rightzhushi = lines[i].find("*/")
            # 找到另一半注释
            while rightzhushi == -1:
                i += 1
                if i >= len(lines):  # 防止有的人只写注释的前一半，不写后一半
                    break
                rightzhushi = lines[i].find("*/")
            i += 1
            continue
        # 如果有删除的修改，记录加一
        if lines[i].startswith("-"):
            # 如果在删除的一行没有分号，说明这不是一条完整的语句，不是我们需要的内容，直接下一行
            if lines[i].find(';') == -1:
                # 如果是没有分号，只有是左大括号才能加一
                if lines[i].find("{") != -1:
                    changecount += 1
                i += 1
                continue
            changecount += 1
        # 什么都没有发生 到下一行
        i += 1
    # 返回修改次数
    return changecount

#生成markdown表格，放到github库中
#path 存放数据的excel表格的路径
def makeTable(path):
    data_excel = xlrd.open_workbook(path)
    table = data_excel.sheets()[0]
    col_name = table.col(colx=0) #用户和库的名称
    col_event = table.col(colx=2) #event use
    col_event_loc = table.col(colx=3) #event use / LOC
    # col_emitcodechurn = table.col(colx=10) #event use code churn
    # col_entirecodechurn = table.col(colx=11) #entire code code churn
    col_star = table.col(colx=5)
    print("hhh")

    f = open("address.txt","w")
    ff = open("message.txt","w")
    #f 是用来存放生成的2915个库的名称和链接
    #ff 是用来存放生成的2915个库的名称、event use   event use/LOC   churn rate entire code    churn rate event use code
    # f.write("|serial number|  user name  | repository name | link |\n|    :---:    |      :---:     |    :---:    |    :--------:    |\n")
    ff.write("|Index|Star| Repository Name | Event Use |Event Use/LOC|\n|    ----    |      ----     |    ----    |    ----    |    ----    |\n")
    for i in range(len(col_name)):
        j = str(col_star[i]).split(':')[-1].split('.')[0]
        star = int(j)
        if star < 5:
            star = 5
        # print(str(col_name[i]).split(" ")[0].split('\'')[-1] + "|" + str(col_name[i]).split(" ")[-1].split('\'')[0])
        # f.write("|" + str(i) + "|" + str(col_name[i]).split(" ")[0].split('\'')[-1] + "|" + str(col_name[i]).split(" ")[-1].split('\'')[0] + "|" + "https://github.com/" + str(col_name[i]).split(" ")[0].split('\'')[-1] + "/" + str(col_name[i]).split(" ")[-1].split('\'')[0] + "\n")
        ff.write("|"+ str(i+1) +  "|" + str(col_name[i]).split(" ")[0].split('\'')[-1] + "/" + str(col_name[i]).split(" ")[-1].split('\'')[0] + "|" +
                str(star) + "|" + str(col_event[i]).split(':')[-1].split('.')[0] + '|' + str(col_event_loc[i]).split(":")[-1] +"\n")



def test(a,b):
    try:
        print(a/b)
    except:
        print("except")
    else :
        print("else")
    print("nothing")








